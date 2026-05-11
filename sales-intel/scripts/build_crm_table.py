#!/usr/bin/env python3
"""
Builds a CRM-ready master table from all sales-intel artifacts.

Inputs (relative to repo root):
  sales-intel/raw/bo_okved_raw.jsonl                — 5326 records (full universe)
  sales-intel/data/bakery-leads-okved-enriched.csv  — top-150 with contacts
  sales-intel/data/bakery-leads-sausage-tested.csv  — top-150 with sausage check

Outputs:
  sales-intel/data/crm-master-q1-2026.csv           — flat CSV (UTF-8 BOM for Excel/Sheets)
  sales-intel/data/crm-master-q1-2026.xlsx          — XLSX with frozen header, autofilter,
                                                      data validation for status/priority/owner

CRM schema (columns):
  status, owner, priority, last_touch, next_action, notes,
  tier, revenue_band, sausage_in_dough, has_contacts,
  company_short, legal_form, inn, ogrn,
  okved_main, okved_name,
  revenue_mln_rub, revenue_period, status_egrul, registration_date,
  region, city, address,
  phones, emails, website,
  evidence_label, evidence_url, evidence_snippet,
  inn_link, source
"""

from __future__ import annotations

import csv
import json
import math
import os
import sys
from datetime import date

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
RAW_JSONL = os.path.join(ROOT, "sales-intel/raw/bo_okved_raw.jsonl")
ENRICHED_CSV = os.path.join(ROOT, "sales-intel/data/bakery-leads-okved-enriched.csv")
SAUSAGE_CSV = os.path.join(ROOT, "sales-intel/data/bakery-leads-sausage-tested.csv")

OUT_CSV = os.path.join(ROOT, "sales-intel/data/crm-master-q1-2026.csv")
OUT_XLSX = os.path.join(ROOT, "sales-intel/data/crm-master-q1-2026.xlsx")

OKOPF_NAMES = {
    12300: "ООО",
    12267: "ИП",
    20107: "АО (непубл.)",
    12247: "ИП",
    20100: "ПАО",
    20109: "ЗАО",
    14200: "ПК",
    15300: "КФХ",
    65243: "Гос.предприятие",
    20608: "АО",
    14100: "ПК",
    20110: "АО",
    20112: "АО",
    71400: "Прочее",
}

OKVED_NAMES = {
    "10.71": "Хлеб и мучные кондитерские (10.71)",
    "10.71.1": "Хлеб, хлебобулочные, кондитерские недлит. хранения (10.71.1)",
    "10.71.2": "Кондитерские изделия недлит. хранения (10.71.2)",
    "10.71.3": "Макаронные изделия (10.71.3)",
    "10.72": "Сухари, печенье, длит. хранения (10.72)",
    "10.72.1": "Сухари, гренки, печенье (10.72.1)",
    "10.72.2": "Пряники и коврижки (10.72.2)",
    "10.72.3": "Прочие длит. хранения (10.72.3)",
}


def okved_top(okved: str) -> str:
    """10.71.2 -> 10.71"""
    if not okved:
        return ""
    parts = okved.split(".")
    if len(parts) >= 2:
        return f"{parts[0]}.{parts[1]}"
    return okved

# Status pipeline values (for data validation in xlsx)
STATUS_VALUES = [
    "new",
    "qualified",
    "contacted",
    "meeting",
    "proposal",
    "negotiation",
    "won",
    "lost",
    "freeze",
    "skip",
]
PRIORITY_VALUES = ["A+", "A", "B", "C", "—"]
OWNER_VALUES = ["", "Ринат", "Менеджер 1", "Менеджер 2"]


def revenue_band(mln: float) -> str:
    if mln >= 10000:
        return "≥10 млрд"
    if mln >= 3000:
        return "3-10 млрд"
    if mln >= 1000:
        return "1-3 млрд"
    if mln >= 300:
        return "300 млн-1 млрд"
    if mln >= 100:
        return "100-300 млн"
    if mln > 0:
        return "<100 млн"
    return "нет данных"


def tier(mln: float) -> str:
    if mln >= 3000:
        return "gigant"
    if mln >= 300:
        return "core"
    if mln >= 100:
        return "growth"
    if mln > 0:
        return "small"
    return "unknown"


def priority(mln: float) -> str:
    if mln >= 1000:
        return "A+"
    if mln >= 300:
        return "A"
    if mln >= 100:
        return "B"
    if mln > 0:
        return "C"
    return "—"


def legal_form(okopf) -> str:
    try:
        return OKOPF_NAMES.get(int(okopf), str(okopf or ""))
    except (TypeError, ValueError):
        return ""


def normalize_name(s: str) -> str:
    if not s:
        return ""
    s = s.replace('\\"', '"').strip()
    return s


def build_index_enriched():
    """Returns dict[inn] -> {phones, emails, sites}."""
    out = {}
    if not os.path.exists(ENRICHED_CSV):
        return out
    with open(ENRICHED_CSV, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            inn = (r.get("inn") or "").strip()
            if not inn:
                continue
            out[inn] = {
                "phones": (r.get("phones") or "").strip(),
                "emails": (r.get("emails") or "").strip(),
                "sites": (r.get("sites") or "").strip(),
            }
    return out


def build_index_sausage():
    """Returns dict[inn] -> {classification, evidence_label, evidence_url, evidence_snippet}."""
    out = {}
    if not os.path.exists(SAUSAGE_CSV):
        return out
    with open(SAUSAGE_CSV, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            inn = (r.get("inn") or "").strip()
            if not inn:
                continue
            cls = (r.get("classification") or "").strip()
            label = (r.get("evidence_label") or "").strip()
            short_label = ""
            if cls.startswith("yes_sausage_in_dough"):
                short_label = "yes"
            elif cls.startswith("yes_hotdog_or_meat_pirozhki"):
                short_label = "probably"
            elif cls in ("no_evidence", "no"):
                short_label = "no"
            elif cls in ("no_website", "fetch_error", "scan_error", "blocked"):
                short_label = "—"
            else:
                short_label = cls or "—"
            out[inn] = {
                "label": short_label,
                "evidence_label": label,
                "evidence_url": (r.get("evidence_url") or "").strip(),
                "evidence_snippet": (r.get("evidence_snippet") or "").strip(),
            }
    return out


def main() -> None:
    enriched = build_index_enriched()
    sausage = build_index_sausage()

    rows = []
    with open(RAW_JSONL, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            inn = (r.get("inn") or "").strip()
            rev_tsd = r.get("revenue_tsd_rub") or 0
            try:
                rev_mln = float(rev_tsd) / 1000.0
            except (TypeError, ValueError):
                rev_mln = 0.0

            enr = enriched.get(inn, {})
            ssg = sausage.get(inn, {})

            okved = (r.get("okved2") or "").strip()
            row = {
                # CRM workflow (manager fills in)
                "status": "new",
                "owner": "",
                "priority": priority(rev_mln),
                "last_touch": "",
                "next_action": "",
                "notes": "",
                # Computed segmentation
                "tier": tier(rev_mln),
                "revenue_band": revenue_band(rev_mln),
                "sausage_in_dough": ssg.get("label", ""),
                "has_contacts": "yes" if (enr.get("phones") or enr.get("emails") or enr.get("sites")) else "",
                # Base data
                "company_short": normalize_name(r.get("name_short") or ""),
                "legal_form": legal_form(r.get("okopf")),
                "inn": inn,
                "ogrn": (r.get("ogrn") or "").strip(),
                "okved_main": okved,
                "okved_top": okved_top(okved),
                "okved_name": OKVED_NAMES.get(okved, OKVED_NAMES.get(okved_top(okved), okved)),
                "revenue_mln_rub": round(rev_mln, 2) if rev_mln else "",
                "revenue_period": (r.get("bfo_period") or "").strip(),
                "status_egrul": (r.get("status_code") or "").strip(),
                "registration_date": (r.get("status_date") or "").strip(),
                # Geography
                "region": (r.get("region_name") or "").strip().title(),
                "city": (r.get("city") or "").strip().title(),
                "address": (r.get("address") or "").strip(),
                # Contacts
                "phones": enr.get("phones", ""),
                "emails": enr.get("emails", ""),
                "website": enr.get("sites", ""),
                # Evidence (sausage scan)
                "evidence_label": ssg.get("evidence_label", ""),
                "evidence_url": ssg.get("evidence_url", ""),
                "evidence_snippet": ssg.get("evidence_snippet", "")[:500],
                # Source links
                "inn_link": f"https://zachestnyibiznes.ru/search?query={inn}" if inn else "",
                "source": "bo.nalog.gov.ru/advanced-search 2024-2025"
                + (" + zachestnyibiznes" if enr else "")
                + (" + site-scan" if ssg else ""),
            }
            rows.append(row)

    # Sort: priority A+, A, B, C, —; then revenue desc
    pri_order = {"A+": 0, "A": 1, "B": 2, "C": 3, "—": 4}
    rows.sort(
        key=lambda x: (
            pri_order.get(x["priority"], 99),
            -(x["revenue_mln_rub"] if isinstance(x["revenue_mln_rub"], (int, float)) else 0),
            x["company_short"],
        )
    )

    columns = list(rows[0].keys()) if rows else []

    # Write CSV with UTF-8 BOM (so Excel/Sheets pick up encoding right)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        w.writerows(rows)
    print(f"[csv] {OUT_CSV}  rows={len(rows)}")

    # Write XLSX with header styling, freeze pane, autofilter, validation
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("openpyxl not installed — skipping xlsx", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(columns)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r[c] for c in columns])

    ws.freeze_panes = "G2"  # freeze CRM workflow columns + header

    # Autofilter on whole used range
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Column widths (approximate)
    width_map = {
        "status": 14, "owner": 14, "priority": 9, "last_touch": 12, "next_action": 12,
        "notes": 30,
        "tier": 10, "revenue_band": 14, "sausage_in_dough": 14, "has_contacts": 12,
        "company_short": 38, "legal_form": 12, "inn": 13, "ogrn": 16,
        "okved_main": 11, "okved_top": 10, "okved_name": 36,
        "revenue_mln_rub": 14, "revenue_period": 9, "status_egrul": 10, "registration_date": 13,
        "region": 22, "city": 18, "address": 40,
        "phones": 24, "emails": 30, "website": 28,
        "evidence_label": 16, "evidence_url": 30, "evidence_snippet": 60,
        "inn_link": 24, "source": 28,
    }
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(col, 15)

    # Data validation for status / priority / owner
    def add_validation(col_name: str, values: list[str]):
        if col_name not in columns:
            return
        col_letter = get_column_letter(columns.index(col_name) + 1)
        # quote-comma list (Excel data validation list)
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
        ws.add_data_validation(dv)

    add_validation("status", STATUS_VALUES)
    add_validation("priority", PRIORITY_VALUES)
    add_validation("owner", OWNER_VALUES)

    # Header row height
    ws.row_dimensions[1].height = 28

    # Write a small help sheet
    help_ws = wb.create_sheet("README")
    help_lines = [
        ["Pepperoni — CRM Master (Q1 2026)", ""],
        ["Generated", date.today().isoformat()],
        ["Total leads", len(rows)],
        ["", ""],
        ["Status pipeline", " → ".join(STATUS_VALUES)],
        ["Priority", "A+ (≥1 млрд) · A (300-1000) · B (100-300) · C (<100) · — (нет БФО)"],
        ["Tier", "gigant ≥3 млрд · core 300-3000 · growth 100-300 · small <100 · unknown"],
        ["sausage_in_dough", "yes / probably / no / — (не проверено или сайт недоступен)"],
        ["", ""],
        ["Source", "ГИР БО (bo.nalog.gov.ru) — ОКВЭД 10.71 + 10.72, БФО 2024+2025"],
        ["Enrichment", "zachestnyibiznes.ru — phones/emails/sites для топ-150"],
        ["Sausage scan", "автоматический crawler по сайту (homepage + catalog)"],
        ["", ""],
        ["Поля для ручного ведения", "status, owner, priority (можно менять), last_touch, next_action, notes"],
        ["Колонки A-F заморожены", "чтобы CRM-поля всегда были видны при горизонт. скролле"],
    ]
    for row in help_lines:
        help_ws.append(row)
    help_ws.column_dimensions["A"].width = 28
    help_ws.column_dimensions["B"].width = 80

    wb.save(OUT_XLSX)
    print(f"[xlsx] {OUT_XLSX}  rows={len(rows)}  cols={len(columns)}")


if __name__ == "__main__":
    main()
